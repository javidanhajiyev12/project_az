from flask import Flask, request, jsonify, send_from_directory, send_file
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image as ExcelImage
from barcode.codex import Code128
from barcode.writer import ImageWriter
from PIL import Image as PILImage
import io, os, base64
from datetime import datetime

app = Flask(__name__)
EXCEL_FILE = "products.xlsx"
PHOTO_DIR = "product_photos"
os.makedirs(PHOTO_DIR, exist_ok=True)

green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
red_fill = PatternFill(start_color='FF7F7F', end_color='FF7F7F', fill_type='solid')

@app.route("/")
def index():
    return send_from_directory('.', 'index.html')

@app.route("/check_barcode", methods=["POST"])
def check_barcode():
    data = request.get_json()
    barcode = str(data.get("barcode", "")).strip()

    if not barcode:
        return jsonify({"status": "error", "message": "❌ Barcode is missing."})

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    found = False
    duplicate_img = False

    for row in ws.iter_rows(min_row=2):
        code = str(row[0].value or "").strip()
        if code == barcode:
            found = True
            product = str(row[1].value or "").strip()

            for img in ws._images:
                if img.anchor._from.row + 1 == row[0].row and img.anchor._from.col == 2:
                    duplicate_img = True

            fill = green_fill if product.upper() != "TAPILMADI" else red_fill
            for cell in row:
                cell.fill = fill

            status = "green" if fill == green_fill else "red"
            message = "✅ Barkod tapıldı və düzgündür." if status == "green" else "❌ Barkod tapıldı, amma məhsul adı 'TAPILMADI'"
            break

    if not found:
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1).value = barcode
        ws.cell(row=new_row, column=2).value = "TAPILMADI"
        ws.cell(row=new_row, column=1).fill = yellow_fill
        ws.cell(row=new_row, column=2).fill = yellow_fill
        status = "yellow"
        message = "➕ Barkod tapılmadı — əlavə olundu."

    if not duplicate_img:
        try:
            buffer = io.BytesIO()
            Code128(barcode, writer=ImageWriter()).write(buffer)
            buffer.seek(0)
            pil = PILImage.open(buffer)
            png = io.BytesIO()
            pil.save(png, format="PNG")
            png.seek(0)

            excel_img = ExcelImage(png)
            excel_img.width = 150
            excel_img.height = 50

            for row in ws.iter_rows(min_row=2):
                if str(row[0].value or "").strip() == barcode:
                    ws.add_image(excel_img, f"C{row[0].row}")
                    break
        except Exception as e:
            print(f"⚠️ Barcode image error: {e}")

    wb.save(EXCEL_FILE)
    return jsonify({"status": status, "message": message})

@app.route("/upload_photo", methods=["POST"])
def upload_photo():
    data = request.get_json()
    barcode = str(data.get("barcode", "")).strip()
    image_data = data.get("image")

    if not barcode or not image_data:
        return jsonify({"message": "❗ Barcode or photo missing"}), 400

    try:
        header, encoded = image_data.split(",", 1)
        img_data = base64.b64decode(encoded)
        img = PILImage.open(io.BytesIO(img_data))

        filename = f"{barcode}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        path = os.path.join(PHOTO_DIR, filename)
        img.save(path)

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        photo_img = ExcelImage(path)
        photo_img.width = 150
        photo_img.height = 100

        for row in ws.iter_rows(min_row=2):
            if str(row[0].value or "").strip() == barcode:
                ws.add_image(photo_img, f"D{row[0].row}")
                break
        else:
            r = ws.max_row + 1
            ws.cell(row=r, column=1).value = barcode
            ws.cell(row=r, column=2).value = "TAPILMADI"
            ws.cell(row=r, column=1).fill = yellow_fill
            ws.cell(row=r, column=2).fill = yellow_fill
            ws.add_image(photo_img, f"D{r}")

        wb.save(EXCEL_FILE)
        return jsonify({"message": f"✅ Photo saved for {barcode}"})
    except Exception as e:
        return jsonify({"message": f"❌ Upload failed: {e}"}), 500

@app.route("/download_excel")
def download_excel():
    return send_file(EXCEL_FILE, as_attachment=True)

if __name__ == "__main__":
    app.run(debug=True)



